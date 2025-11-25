import re, math
from itertools import combinations
from typing import Optional, List, Dict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os

# ---------- Config ----------
T_SHEETS    = {"T1", "T2", "T3", "T4"}
FINAL_COLS  = ["pregunta", "concepto", "n", "porcentaje", "DS"]
START_COL   = 4       # buscar bloques desde columna D en adelante
BASE_MIN    = 50      # regla del negocio
MAX_SET     = 4
ADD_ARROW   = True    # anteponer "▲ " al DS (Concatenado) cuando haya texto

# PDP nuevo
PDP_BLOCKS = [
      {"block": "FILTRO Y PERFILAMIENTO",     "sheet": "T2", "range": ("P1","P6"),   "split": False, "flags": []},
      {"block": "CONSUMO DE LA CATEGORÍA",    "sheet": "T2", "range": ("P7","P13"),  "split": False, "flags": []},
      {"block": "POSICIONAMIENTO",            "sheet": "T2", "range": ("P14","P18"), "split": False, "flags": []},
      {"block": "EVALUACION DE ANAQUEL",      "sheet": "T2", "range": ("P19","P26"), "split": True,  "flags": ["a","b"]},
      {"block": "EVALUACION DE CONCEPTO",     "sheet": "T2", "range": ("P27","P58"), "split": True,  "flags": ["a","b"]},
]

PDP_CONFIG = (False, "T2")

# ---------- Normalización ----------
def _norm(s) -> str:
    if s is None:
        return ""
    t = str(s).replace("\u00A0", " ").strip()
    while "  " in t:
        t = t.replace("  ", " ")
    return t

def _is_base_label(val) -> bool:
    return _norm(val).casefold() == "base"

# ---------- Números ----------
def _to_number_pct(x):
    if x is None:
        return None
    if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x)):
        return float(x)
    s = str(x).strip().replace("%", "").replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except:
        return None

def _parse_float_safe(val):
    try:
        s = str(val).strip().replace("%", "").replace(",", ".")
        return float(s)
    except:
        return None

# ---------- Merges ----------
def _merged_range(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None

def _merged_top_row(ws, row: int, col: int) -> int:
    rng = _merged_range(ws, row, col)
    return rng.min_row if rng else row

def _merged_bottom_row(ws, row: int, col: int) -> int:
    rng = _merged_range(ws, row, col)
    return rng.max_row if rng else row

# ---------- Anchajes ----------
def _find_first_base_in_col(ws, col_idx: int) -> Optional[int]:
    for r in range(1, ws.max_row + 1):
        if _is_base_label(ws.cell(r, col_idx).value):
            return r
    return None

def _next_base_row_in_B(ws, start_row_exclusive: int) -> Optional[int]:
    for r in range(start_row_exclusive + 1, ws.max_row + 1):
        if _is_base_label(ws.cell(r, 2).value):
            return r
    return None

# ---------- Preguntas ----------
_PNUM_RE = re.compile(r'^p(\d+)', re.I)
_SUB_RE  = re.compile(r'(?i)^p(\d+)([a-z])')
_NUMERIC_CONCEPT_RE = re.compile(r'^\d+(\.\d+)?$')

def _starts_with_pnum(s: str) -> bool:
    return bool(_PNUM_RE.match(_norm(s)))

def _get_pnum(s: str) -> Optional[int]:
    m = _PNUM_RE.match(_norm(s))
    return int(m.group(1)) if m else None

def _extract_subclass(pregunta: str) -> Optional[str]:
    s = _norm(pregunta)
    m = _SUB_RE.match(s)
    if not m:
        return None
    letter = m.group(2).lower()
    if letter == 'r':
        tail = s[m.end(): m.end() + 2].lower()
        if tail.startswith('ec'):
            return None
        return None
    return letter if letter in {'a', 'b'} else None

def _is_numeric_concept(concepto: str) -> bool:
    return bool(_NUMERIC_CONCEPT_RE.match(_norm(concepto)))

# ---------- Bloques horizontales ----------
def _find_header_row_and_bases(ws):
    best_row, best_cols = None, []
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(1, max_row + 1):
        cols = []
        for c in range(START_COL, max_col + 1):
            if _is_base_label(ws.cell(r, c).value):
                cols.append(c)
        if len(cols) > len(best_cols):
            best_row, best_cols = r, cols
    return best_row, sorted(best_cols)

def _scan_horizontal_blocks_dynamic(ws):
    header_row, base_cols = _find_header_row_and_bases(ws)
    blocks = []
    
    # GRID: Si no hay columnas Base, buscar primer texto en columnas
    if not header_row or not base_cols:
        # Buscar Base en columna B para encontrar el header
        base_row_B = _find_first_base_in_col(ws, 2)
        if not base_row_B:
            return None, blocks
        
        # El header suele estar arriba de la fila Base
        header_row = base_row_B - 2
        max_col = ws.max_column
        
        # Crear un solo bloque con todas las columnas desde START_COL
        cols, names, letters = [], [], []
        for c in range(START_COL, max_col + 1):
            nm = _norm(ws.cell(header_row, c).value)
            if nm:  # Si hay contenido
                lt = _norm(ws.cell(header_row + 1, c).value).lower()
                cols.append(c)
                names.append(nm)
                letters.append(lt)
        
        if cols:
            blocks.append({"cols": cols, "names": names, "letters": letters})
        return header_row, blocks

    max_col = ws.max_column
    for i, start in enumerate(base_cols):
        end = (base_cols[i + 1] - 1) if (i + 1 < len(base_cols)) else max_col

        cols, names, letters = [], [], []
        for c in range(start, end + 1):
            nm = _norm(ws.cell(header_row, c).value)
            lt = _norm(ws.cell(header_row + 1, c).value).lower()
            if _is_base_label(nm):
                continue
            cols.append(c)
            names.append(nm)
            letters.append(lt)

        if cols:
            blocks.append({"cols": cols, "names": names, "letters": letters})

    return header_row, blocks

# ---------- Ganadores ----------
def _winners_label_from_names(winners_mask, names):
    lab = [names[i] for i, w in enumerate(winners_mask) if w and names[i]]
    return ", ".join(lab) if lab else "(Sin ganadores)"

def _compute_winners_for_concept_block(ws, topR: int, cols: list, letters: list, names: list, base_min: float = BASE_MIN) -> str:
    S = len(cols)

    # Si solo hay un segmento → no hay ganador
    if S <= 1:
        return "(Sin ganadores)"

    print(f"\n=== DEBUG: topR={topR}, cols={cols}, letters={letters}, names={names} ===")
    
    # Leer DS de todas las columnas (sin filtrar por base)
    beats = [[False] * S for _ in range(S)]
    for i in range(S):
        ds_txt = _norm(ws.cell(topR + 2, cols[i]).value).lower()
        print(f"  Col {cols[i]} ({names[i]}): DS='{ds_txt}'")
        
        for j in range(S):
            if i == j:
                continue
            lj = letters[j] if j < len(letters) else ""
            if lj and lj in ds_txt:
                beats[i][j] = True

    # Buscar ganador único: gana a todos los demás
    winners = [False] * S
    for i in range(S):
        if all(i == j or beats[i][j] for j in range(S)):
            winners[i] = True
    
    if any(winners):
        return _winners_label_from_names(winners, names)

    # Buscar ganadores múltiples (empate + dominan a los externos)
    max_k = min(MAX_SET, max(1, S - 1))
    for k in range(2, max_k + 1):
        for combo in combinations(range(S), k):
            ok = True
            # Consistencia interna: no se ganan entre ellos
            for a in combo:
                for b in combo:
                    if a != b and beats[a][b]:
                        ok = False
                        break
                if not ok:
                    break
            if not ok:
                continue
            
            # Dominancia externa: todos ganan a todos los externos
            for ext in range(S):
                if ext in combo:
                    continue
                for a in combo:
                    if not beats[a][ext]:
                        ok = False
                        break
                if not ok:
                    break
            
            if ok:
                winners = [i in combo for i in range(S)]
                return _winners_label_from_names(winners, names)

    return "(Sin ganadores)"

# ---------- DS CONCATEANDO ----------
def build_ds_concat_map_for_ws(ws, base_min: float = BASE_MIN) -> dict:
    ds_concat = {}
    header_row, blocks = _scan_horizontal_blocks_dynamic(ws)
    if not header_row or not blocks:
        return ds_concat

    first_base_D = _find_first_base_in_col(ws, 4)
    if first_base_D is None:
        # GRID: Buscar Base en columna B
        first_base_B = _find_first_base_in_col(ws, 2)
        if first_base_B is None:
            return ds_concat
        probe_row = first_base_B
    else:
        probe_row = first_base_D + 2

    start_block_row = None
    for r in range(probe_row, ws.max_row + 1):
        if _is_base_label(ws.cell(r, 2).value):
            start_block_row = r
            break
    if start_block_row is None:
        return ds_concat

    current_base_row = start_block_row
    while current_base_row and current_base_row <= ws.max_row:
        next_base_row = _next_base_row_in_B(ws, current_base_row)
        end_row_exclusive = next_base_row if next_base_row else (ws.max_row + 1)

        r = current_base_row + 1
        while r < end_row_exclusive:
            concepto = _norm(ws.cell(r, 2).value)
            if not concepto or _is_base_label(concepto):
                r += 1
                continue

            topR = _merged_top_row(ws, r, 2)

            seen, ordered = set(), []
            for blk in blocks:
                lab = _compute_winners_for_concept_block(ws, topR, blk["cols"], blk["letters"], blk["names"], base_min=base_min)
                if lab and lab != "(Sin ganadores)":
                    for token in [t.strip() for t in lab.split(",") if t.strip()]:
                        if token not in seen:
                            seen.add(token)
                            ordered.append(token)

            concat_txt = ", ".join(ordered) if ordered else ""
            if concat_txt and ADD_ARROW:
                concat_txt = f"▲ {concat_txt}"

            ds_concat[topR] = concat_txt

            r = _merged_bottom_row(ws, r, 2) + 1 if _merged_range(ws, r, 2) else (r + 3)

        current_base_row = next_base_row

    return ds_concat

# ---------- Extracción filas ----------
def extract_rows_from_ws(ws, base_min: float = BASE_MIN) -> List[Dict]:
    rows = []
    ds_concat_map = build_ds_concat_map_for_ws(ws, base_min=base_min)

    first_base_D = _find_first_base_in_col(ws, 4)
    is_grid = (first_base_D is None)
    
    if is_grid:
        # GRID: Buscar Base en columna B
        first_base_B = _find_first_base_in_col(ws, 2)
        if first_base_B is None:
            return rows
        probe_row = first_base_B
    else:
        probe_row = first_base_D + 2
    if probe_row > ws.max_row:
        return rows

    start_block_row = None
    for r in range(probe_row, ws.max_row + 1):
        if _is_base_label(ws.cell(r, 2).value):
            start_block_row = r
            break
    if start_block_row is None:
        return rows

    current_base_row = start_block_row
    while current_base_row and current_base_row <= ws.max_row:
        next_base_row = _next_base_row_in_B(ws, current_base_row)
        end_row_exclusive = next_base_row if next_base_row else (ws.max_row + 1)

        pregunta = _norm(ws.cell(current_base_row, 1).value)

        r = current_base_row + 1
        while r < end_row_exclusive:
            concepto = _norm(ws.cell(r, 2).value)
            if not concepto or _is_base_label(concepto):
                r += 1
                continue

            topR = _merged_top_row(ws, r, 2)

            ds_val = ds_concat_map.get(topR, "")
            
            if is_grid:
                # GRID: Solo pregunta, concepto, DS
                rows.append({
                    "pregunta": pregunta,
                    "concepto": concepto,
                    "DS": ds_val
                })
            else:
                # NORMAL: pregunta, concepto, n, porcentaje, DS
                n_val = _to_number_pct(ws.cell(r, 4).value)
                n_val = float(n_val) if n_val is not None else 0.0

                pr = r + 1
                if pr >= end_row_exclusive:
                    porcentaje = 0.0
                else:
                    parsed = _to_number_pct(ws.cell(pr, 4).value)
                    porcentaje = float(parsed) if parsed is not None else 0.0

                rows.append({
                    "pregunta": pregunta,
                    "concepto": concepto,
                    "n": n_val,
                    "porcentaje": porcentaje,
                    "DS": ds_val
                })

            r = _merged_bottom_row(ws, r, 2) + 1 if _merged_range(ws, r, 2) else (r + 3)

        current_base_row = next_base_row

    return rows

# ---------- Reglas globales ----------
def apply_global_rules(rows: List[Dict]) -> List[Dict]:
    filtered = []
    for row in rows:
        p = row.get("pregunta", "")
        c = row.get("concepto", "")
        # Procesar todas las preguntas, no solo las que empiezan con P
        # (comentado el filtro restrictivo)
        # if not _starts_with_pnum(p):
        #     continue
        c_norm = _norm(c)
        if c_norm.lower() == "media":
            continue
        if _is_numeric_concept(c_norm):
            continue
        if "porcentaje" not in row or row["porcentaje"] is None:
            row["porcentaje"] = 0.0
        if "n" not in row or row["n"] is None:
            row["n"] = 0.0
        filtered.append(row)

    if not filtered:
        return []

    has_t2b = {}
    for row in filtered:
        p = row["pregunta"]
        c = _norm(row["concepto"])
        if c.upper() == "T2B" or c == "10-9":
            has_t2b[p] = True

    reduced = []
    for row in filtered:
        p = row["pregunta"]
        c = _norm(row["concepto"])
        if has_t2b.get(p, False):
            if c.upper() == "T2B" or c == "10-9":
                r2 = row.copy()
                r2["concepto"] = "T2B"
                reduced.append(r2)
        else:
            reduced.append(row)

    return reduced

# ---------- Excel Styling ----------
def apply_styles(writer, sheet_name, df):
    """
    Aplica formato profesional con branding de Atlantia a la hoja de Excel generada:
    - Encabezados con color corporativo Atlantia (#482F91) y negrita.
    - Filtros automáticos.
    - Paneles inmovilizados.
    - Ancho de columnas inteligente.
    - Formato de % para columnas de porcentaje.
    """
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]
    (max_row, max_col) = df.shape

    # --- Definición de Formatos ---
    # Encabezado: Color primario Atlantia (#482F91), texto blanco, negrita, centrado
    header_fmt = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'vcenter',
        'align': 'center',
        'fg_color': '#482F91',  # Color primario Atlantia
        'font_color': 'white',
        'border': 1
    })

    # Texto general: Alineación vertical centrada, ajuste de texto si es largo
    text_fmt = workbook.add_format({
        'valign': 'vcenter',
        'text_wrap': True
    })
    
    # Números centrados
    num_fmt = workbook.add_format({
        'valign': 'vcenter',
        'align': 'center'
    })

    # Porcentajes: Sin decimales (o usa '0.0%' para 1 decimal)
    pct_fmt = workbook.add_format({
        'num_format': '0%',
        'valign': 'vcenter',
        'align': 'center'
    })

    # Formato para ganadores (DS con ▲): Fondo morado claro Atlantia
    winner_fmt = workbook.add_format({
        'bg_color': '#E8DAFF',      # Morado claro (tinte del color secundario Atlantia)
        'font_color': '#482F91',    # Color primario Atlantia
        'valign': 'vcenter',
        'text_wrap': True,
        'bold': True
    })

    # --- Aplicar formato a los encabezados ---
    # Sobreescribimos los encabezados de Pandas para ponerles nuestro estilo
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_fmt)

    # --- Configuración de la Hoja ---
    # Agregar Autofiltros en los encabezados
    worksheet.autofilter(0, 0, max_row, max_col - 1)
    
    # Inmovilizar la primera fila (encabezados)
    worksheet.freeze_panes(1, 0)

    # --- Ajuste de Ancho de Columnas y Formato de Celdas ---
    col_indices = {}  # Mapeo de nombre de columna a índice
    
    for i, col in enumerate(df.columns):
        # Lógica para ancho de columnas
        col_name = str(col).lower()
        col_indices[col_name] = i
        
        # Columnas de Texto grandes (Pregunta, Concepto)
        if "pregunta" in col_name or "concepto" in col_name:
            worksheet.set_column(i, i, 40, text_fmt)  # Ancho 40
        
        # Columnas de Porcentaje
        elif "porcentaje" in col_name or "%" in col_name:
            worksheet.set_column(i, i, 12, pct_fmt)  # Ancho 12 y formato %
            
        # Columnas numéricas cortas (n, Bloques)
        elif "n" == col_name or "bloque" in col_name:
             worksheet.set_column(i, i, 10, num_fmt)
             
        # Columna DS / Concatenado (puede ser larga pero no tanto como pregunta)
        elif "ds" in col_name or "concatenado" in col_name:
            worksheet.set_column(i, i, 25, text_fmt)
            
        # Default para cualquier otra cosa
        else:
            worksheet.set_column(i, i, 15, text_fmt)
    
    # --- Formato Condicional para Ganadores (DS/Concatenado) ---
    # Resaltar celdas que contienen el símbolo de ganador "▲"
    if 'ds' in col_indices:
        ds_col = col_indices['ds']
        worksheet.conditional_format(1, ds_col, max_row, ds_col, {
            'type': 'text',
            'criteria': 'containing',
            'value': '▲',
            'format': winner_fmt
        })
    
    if 'concatenado' in col_indices:
        concat_col = col_indices['concatenado']
        worksheet.conditional_format(1, concat_col, max_row, concat_col, {
            'type': 'text',
            'criteria': 'containing',
            'value': '▲',
            'format': winner_fmt
        })

# ---------- PDP utils ----------
def filter_rows_by_range(rows: List[Dict], p_ini: str, p_fin: str) -> List[Dict]:
    n_ini = _get_pnum(p_ini)
    n_fin = _get_pnum(p_fin)
    if n_ini is None or n_fin is None:
        return []
    lo, hi = min(n_ini, n_fin), max(n_ini, n_fin)
    out = []
    for row in rows:
        n = _get_pnum(row.get("pregunta", ""))
        if n is None:
            continue
        if lo <= n <= hi:
            out.append(row)
    return out

def split_by_subclass(rows: List[Dict]) -> Dict[str, List[Dict]]:
    general, A, B = [], [], []
    for row in rows:
        sub = _extract_subclass(row.get("pregunta", ""))
        pack = {k: row[k] for k in FINAL_COLS}
        if sub == "a":
            A.append(pack)
        elif sub == "b":
            B.append(pack)
        else:
            general.append(pack)
    return {"general": general, "a": A, "b": B}

# ---------- Ejecutar ----------
def process_workbook_by_pdp(
    input_file,
    segment_with_pdp: Optional[bool] = None,
    general_sheet: Optional[str] = None,
    base_min: float = BASE_MIN
) -> str:

    cfg_segment, cfg_sheet = PDP_CONFIG
    if segment_with_pdp is None:
        segment_with_pdp = cfg_segment
    if general_sheet is None:
        general_sheet = cfg_sheet

    wb_in = load_workbook(input_file)
    
    # Create temp file for output
    fd, out_path = tempfile.mkstemp(suffix="_BY_BLOCKS.xlsx")
    os.close(fd)
    
    writer = pd.ExcelWriter(out_path, engine="xlsxwriter")

    # Procesar TODAS las hojas (no solo T1-T4)
    ws_map = {name: wb_in[name] for name in wb_in.sheetnames}

    extracted = {}
    for tname, ws in ws_map.items():
        rows = extract_rows_from_ws(ws, base_min=base_min)
        # Excluir filas con base menor al umbral (excepto Grid donde no hay 'n')
        rows = [r for r in rows if ("n" not in r) or (r.get("n", 0) >= base_min)]
        rows = apply_global_rules(rows)
        extracted[tname] = rows

    sheet_key = None
    for tname in extracted.keys():
        if tname.strip().upper() == general_sheet.strip().upper():
            sheet_key = tname
            break

    if sheet_key is not None:
        all_rows = extracted.get(sheet_key, [])
        if all_rows:
            # Detectar Grid: si no tiene 'n', es Grid
            is_grid = "n" not in all_rows[0]
            cols = ["pregunta", "concepto", "DS"] if is_grid else FINAL_COLS
            
            df_gen = pd.DataFrame(all_rows, columns=cols)
            s_name = "GENERAL"
            df_gen.to_excel(writer, sheet_name=s_name, index=False)
            apply_styles(writer, s_name, df_gen)

    if segment_with_pdp:
        for blk in PDP_BLOCKS:
            blk_name = blk["block"]
            tname    = blk["sheet"]
            p_ini, p_fin = blk["range"]
            do_split  = blk["split"]
            flags     = blk["flags"]

            rows_sheet = extracted.get(tname, [])
            rows_block = filter_rows_by_range(rows_sheet, p_ini, p_fin)

            if not rows_block:
                continue

            if not do_split:
                df_blk = pd.DataFrame(rows_block, columns=FINAL_COLS)
                s_name = blk_name[:31]
                df_blk.to_excel(writer, sheet_name=s_name, index=False)
                apply_styles(writer, s_name, df_blk)
            else:
                parts = split_by_subclass(rows_block)

                df_gral = pd.DataFrame(parts["general"], columns=FINAL_COLS)
                s_name_gral = blk_name[:31]
                df_gral.to_excel(writer, sheet_name=s_name_gral, index=False)
                apply_styles(writer, s_name_gral, df_gral)

                for flg in flags:
                    sub_rows = parts.get(flg, [])
                    if sub_rows:
                        df_sub = pd.DataFrame(sub_rows, columns=FINAL_COLS)
                        s_name_sub = f"{blk_name}_{flg}"[:31]
                        df_sub.to_excel(writer, sheet_name=s_name_sub, index=False)
                        apply_styles(writer, s_name_sub, df_sub)

    writer.close()
    return out_path


# ===== Celda 3: Tabla DS por bloque + modo DS ONLY =====

def build_ds_table_for_ws(ws, base_min: float = BASE_MIN):
    """
    Genera tabla a nivel:
      - pregunta
      - concepto
      - Bloque 1..k  (ganador por bloque, o '(Sin ganadores)')
      - Concatenado  (todos los ganadores de izquierda a derecha, sin duplicar)
    Usa la misma lógica de bloques y DS que build_ds_concat_map_for_ws.
    """
    results = []

    header_row, blocks = _scan_horizontal_blocks_dynamic(ws)
    if not header_row or not blocks:
        return results, 0

    # ancla vertical: primera 'Base' en D; +2 filas; primera 'Base' en B
    first_base_D = _find_first_base_in_col(ws, 4)
    if first_base_D is None:
        return results, 0
    probe_row = first_base_D + 2

    # buscar inicio del primer bloque vertical (Base en B)
    start_block_row = None
    for r in range(probe_row, ws.max_row + 1):
        if _is_base_label(ws.cell(r, 2).value):
            start_block_row = r
            break
    if start_block_row is None:
        return results, 0

    current_base_row = start_block_row

    while current_base_row and current_base_row <= ws.max_row:
        next_base_row = _next_base_row_in_B(ws, current_base_row)
        end_row_exclusive = next_base_row if next_base_row else (ws.max_row + 1)

        pregunta = _norm(ws.cell(current_base_row, 1).value)  # col A

        r = current_base_row + 1
        while r < end_row_exclusive:
            concepto = _norm(ws.cell(r, 2).value)  # col B
            if not concepto or _is_base_label(concepto):
                r += 1
                continue

            topR = _merged_top_row(ws, r, 2)

            per_block_labels = []
            seen, ordered = set(), []

            # recorrer bloques izquierda → derecha
            for blk in blocks:
                lab = _compute_winners_for_concept_block(
                    ws,
                    topR,
                    blk["cols"],
                    blk["letters"],
                    blk["names"],
                    base_min=base_min
                )
                per_block_labels.append(lab)

                if lab and lab != "(Sin ganadores)":
                    for token in [t.strip() for t in lab.split(",") if t.strip()]:
                        if token not in seen:
                            seen.add(token)
                            ordered.append(token)

            concat_txt = ", ".join(ordered) if ordered else ""
            if concat_txt and ADD_ARROW:
                concat_txt = f"▲ {concat_txt}"

            row = {
                "pregunta": pregunta,
                "concepto": concepto,
            }

            # Bloque 1, Bloque 2, ...
            for idx, lab in enumerate(per_block_labels, start=1):
                row[f"Bloque {idx}"] = lab

            row["Concatenado"] = concat_txt

            results.append(row)

            # saltos por merges como en extract_rows_from_ws
            r = _merged_bottom_row(ws, r, 2) + 1 if _merged_range(ws, r, 2) else (r + 3)

        current_base_row = next_base_row

    return results, len(blocks)


def process_workbook_ds_only(input_file, sheet_name: str, base_min: float = BASE_MIN) -> str:
    """
    Procesa SOLO una hoja (sheet_name) y genera:
      - Archivo *_DS_ONLY_<sheet>.xlsx
      - Hoja 'DS' con columnas:
        pregunta, concepto, Bloque 1..k, Concatenado

    Ideal para sacar tablas como:
      Concepto | Bloque 1 | Bloque 2 | Bloque 3 | Concatenado
    """
    wb_in = load_workbook(input_file)
    if sheet_name not in wb_in.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el libro.")

    ws = wb_in[sheet_name]
    rows, n_blocks = build_ds_table_for_ws(ws, base_min=base_min)

    if not rows:
        raise ValueError(f"No se encontraron filas de conceptos en la hoja '{sheet_name}'.")

    df = pd.DataFrame(rows)

    # Create temp file for output
    fd, out_path = tempfile.mkstemp(suffix=f"_DS_ONLY_{sheet_name}.xlsx")
    os.close(fd)

    with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
        s_name = "DS"
        df.to_excel(writer, sheet_name=s_name, index=False)
        apply_styles(writer, s_name, df)

    return out_path
